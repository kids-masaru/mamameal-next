from flask import Flask, request, send_file, jsonify
import os
import io
import json
import pandas as pd
from openpyxl import load_workbook
import google.generativeai as genai
from utils import load_master_csv, match_bento_data, safe_write_df, paste_dataframe_to_sheet
import zipfile

app = Flask(__name__)

# Configure Gemini
# Note: In Vercel, set GOOGLE_API_KEY in environment variables
if "GOOGLE_API_KEY" in os.environ:
    genai.configure(api_key=os.environ["GOOGLE_API_KEY"])

ASSETS_DIR = os.path.join(os.path.dirname(__file__), 'assets')

@app.route('/api/process', methods=['POST'])
def process_pdf():
    if 'file' not in request.files:
        return jsonify({"error": "No file part"}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No selected file"}), 400

    try:
        # Read file bytes
        pdf_bytes = file.read()
        
        # Get model from request, default to gemini-2.5-flash
        model_name = request.form.get('model', 'gemini-2.5-flash')

        # Call Gemini API
        model = genai.GenerativeModel(model_name)
        prompt = """
        Analyze this PDF document which contains bento (lunch box) orders and client counts.
        Extract the following information and return it as a JSON object:

        1. "bento_items": A list of all bento items found. For each item, extract the "name" (text).
        2. "clients": A list of clients. For each client, extract:
           - "name": Client name (e.g., specific nursery or school name).
           - "student_meals": A list of numbers representing student meal counts (usually up to 3 numbers).
           - "teacher_meals": A list of numbers representing teacher meal counts (usually up to 2 numbers).
        
        Return ONLY valid JSON. Do not include markdown formatting.
        """
        
        response = model.generate_content([
            {"mime_type": "application/pdf", "data": pdf_bytes},
            prompt
        ])
        
        # Clean response text (remove markdown code blocks if present)
        text = response.text.strip()
        if text.startswith("```json"):
            text = text[7:]
        if text.endswith("```"):
            text = text[:-3]
        
        data = json.loads(text)
        
        # Process Data
        bento_items = data.get("bento_items", [])
        clients = data.get("clients", [])
        
        # Load Masters
        df_product_master = load_master_csv(ASSETS_DIR, "商品マスタ")
        df_customer_master = load_master_csv(ASSETS_DIR, "得意先マスタ")
        
        # Match Bento Data
        bento_names = [item["name"] for item in bento_items]
        matched_bento = match_bento_data(bento_names, df_product_master)
        df_bento_sheet = pd.DataFrame(matched_bento, columns=['商品予定名', 'パン箱入数', '売価単価', '弁当区分'])
        
        # Process Client Data
        client_rows = []
        for client in clients:
            s_meals = client.get("student_meals", [])
            t_meals = client.get("teacher_meals", [])
            row = {
                'クライアント名': client.get("name", ""),
                '園児の給食の数1': s_meals[0] if len(s_meals) > 0 else '',
                '園児の給食の数2': s_meals[1] if len(s_meals) > 1 else '',
                '園児の給食の数3': s_meals[2] if len(s_meals) > 2 else '',
                '先生の給食の数1': t_meals[0] if len(t_meals) > 0 else '',
                '先生の給食の数2': t_meals[1] if len(t_meals) > 1 else ''
            }
            client_rows.append(row)
        df_client_sheet = pd.DataFrame(client_rows)

        # Load Templates
        template_path = os.path.join(ASSETS_DIR, "template.xlsm")
        nouhinsyo_path = os.path.join(ASSETS_DIR, "nouhinsyo.xlsx")
        
        template_wb = load_workbook(template_path, keep_vba=True)
        nouhinsyo_wb = load_workbook(nouhinsyo_path)
        
        # Write to Template
        if not df_product_master.empty and "商品マスタ" in template_wb.sheetnames:
             # Clear and write (simplified)
             paste_dataframe_to_sheet(template_wb["商品マスタ"], df_product_master)
        
        if not df_customer_master.empty and "得意先マスタ" in template_wb.sheetnames:
             paste_dataframe_to_sheet(template_wb["得意先マスタ"], df_customer_master)

        if not df_bento_sheet.empty and "注文弁当の抽出" in template_wb.sheetnames:
            safe_write_df(template_wb["注文弁当の抽出"], df_bento_sheet)
            
        if not df_client_sheet.empty and "クライアント抽出" in template_wb.sheetnames:
            safe_write_df(template_wb["クライアント抽出"], df_client_sheet)

        # Write to Nouhinsyo (Data Only)
        # Logic for mapping product names for Nouhinsyo
        df_bento_for_nouhin = None
        if not df_bento_sheet.empty:
            master_df = df_product_master.copy()
            if not master_df.empty and '商品名' in master_df.columns:
                master_map = master_df.drop_duplicates(subset=['商品予定名']).set_index('商品予定名')['商品名'].to_dict()
                df_bento_for_nouhin = df_bento_sheet.copy()
                df_bento_for_nouhin['商品名'] = df_bento_for_nouhin['商品予定名'].map(master_map)
                df_bento_for_nouhin = df_bento_for_nouhin[['商品予定名', 'パン箱入数', '商品名']]

        if df_bento_for_nouhin is not None and "注文弁当の抽出" in nouhinsyo_wb.sheetnames:
            safe_write_df(nouhinsyo_wb["注文弁当の抽出"], df_bento_for_nouhin)
            
        if not df_client_sheet.empty and "クライアント抽出" in nouhinsyo_wb.sheetnames:
            safe_write_df(nouhinsyo_wb["クライアント抽出"], df_client_sheet)
            
        if not df_customer_master.empty and "得意先マスタ" in nouhinsyo_wb.sheetnames:
            safe_write_df(nouhinsyo_wb["得意先マスタ"], df_customer_master)

        # Save to Bytes
        output_macro = io.BytesIO()
        template_wb.save(output_macro)
        output_macro.seek(0)
        
        output_data = io.BytesIO()
        nouhinsyo_wb.save(output_data)
        output_data.seek(0)
        
        # Create Zip
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w') as zf:
            zf.writestr('数出表.xlsm', output_macro.getvalue())
            zf.writestr('納品書.xlsx', output_data.getvalue())
        zip_buffer.seek(0)
        
        return send_file(
            zip_buffer,
            mimetype='application/zip',
            as_attachment=True,
            download_name='results.zip'
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True)
