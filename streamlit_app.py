import streamlit as st
import os
import sys
import json
import base64
import io
import logging
from pathlib import Path

# ========================================================
# CRITICAL: Suppress pdfminer logging BEFORE any imports
# This must happen at the very start to prevent
# excessive log output that can hang Streamlit Cloud
# ========================================================
logging.getLogger('pdfminer').setLevel(logging.ERROR)
logging.getLogger('pdfplumber').setLevel(logging.ERROR)
logging.getLogger('pdfminer.pdfpage').setLevel(logging.ERROR)
logging.getLogger('pdfminer.pdfinterp').setLevel(logging.ERROR)
logging.getLogger('pdfminer.converter').setLevel(logging.ERROR)
logging.getLogger('pdfminer.pdfdocument').setLevel(logging.ERROR)

import pandas as pd
import google.generativeai as genai
from openpyxl import load_workbook, Workbook
from dotenv import load_dotenv
import unicodedata
import glob

# Robust path resolution for Streamlit Cloud compatibility
APP_DIR = Path(__file__).parent.resolve()
sys.path.insert(0, str(APP_DIR))

from api.pdf_utils import (
    safe_write_df, pdf_to_excel_data_for_paste_sheet, extract_table_from_pdf_for_bento,
    find_correct_anchor_for_bento, extract_bento_range_for_bento, match_bento_data, 
    extract_detailed_client_info_from_pdf, export_detailed_client_data_to_dataframe,
    paste_dataframe_to_sheet
)

# Load environment variables
load_dotenv()

# Configure page
icon_path = os.path.join("static", "icons", "android-chrome-192.png")
page_icon = icon_path if os.path.exists(icon_path) else "🍱"
try:
    if not os.path.exists(icon_path):
        page_icon = "🍱"
except:
    page_icon = "🍱"

st.set_page_config(
    page_title="ママミール業務ツール",
    page_icon=page_icon,
    layout="wide"
)

# Custom CSS & SVGs
st.markdown("""
<style>
    .stButton>button {
        width: 100%;
        background: linear-gradient(to right, #f97316, #ea580c);
        color: white;
        font-weight: bold;
        border: none;
        padding: 0.5rem 1rem;
        border-radius: 0.5rem;
        transition: all 0.3s ease;
    }
    .stButton>button:hover {
        background: linear-gradient(to right, #ea580c, #c2410c);
        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1), 0 2px 4px -1px rgba(0, 0, 0, 0.06);
        transform: translateY(-1px);
    }
    .success-box {
        padding: 1rem;
        background-color: #f0fdf4;
        border: 1px solid #bbf7d0;
        border-radius: 0.5rem;
        color: #15803d;
        margin-bottom: 1rem;
    }
    .main-header {
        font-size: 2.2rem;
        font-weight: 800;
        color: #1e293b;
        margin-bottom: 0.5rem;
        display: flex;
        align-items: center;
        gap: 0.5rem;
    }
    .sub-header {
        font-size: 1.5rem;
        font-weight: 600;
        color: #334155;
        margin-top: 1.5rem;
        margin-bottom: 1rem;
        display: flex;
        align-items: center;
        gap: 0.5rem;
    }
    .card {
        background-color: white;
        padding: 1.5rem;
        border-radius: 0.75rem;
        box-shadow: 0 1px 3px 0 rgba(0, 0, 0, 0.1), 0 1px 2px 0 rgba(0, 0, 0, 0.06);
        margin-bottom: 1rem;
        border: 1px solid #e2e8f0;
    }
    .stTabs [data-baseweb="tab-list"] {
        gap: 2px;
    }
    .stTabs [data-baseweb="tab"] {
        height: 50px;
        white-space: pre-wrap;
        background-color: #f1f5f9;
        border-radius: 4px 4px 0 0;
        gap: 1px;
        padding-top: 10px;
        padding-bottom: 10px;
    }
    .stTabs [aria-selected="true"] {
        background-color: #fff;
        border-bottom: 2px solid #f97316;
    }
</style>
""", unsafe_allow_html=True)

# SVG Icons
ICON_DOC = """<svg xmlns="http://www.w3.org/2000/svg" width="24" height="24" viewBox="0 0 24 24" fill="none" stroke="#f97316" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"></path><polyline points="14 2 14 8 20 8"></polyline><line x1="16" y1="13" x2="8" y2="13"></line><line x1="16" y1="17" x2="8" y2="17"></line><polyline points="10 9 9 9 8 9"></polyline></svg>"""
ICON_TAG = """<svg xmlns="http://www.w3.org/2000/svg" width="24" height="24" viewBox="0 0 24 24" fill="none" stroke="#f97316" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M20.59 13.41l-7.17 7.17a2 2 0 0 1-2.83 0L2 12V2h10l8.59 8.59a2 2 0 0 1 0 2.82z"></path><line x1="7" y1="7" x2="7.01" y2="7"></line></svg>"""
ICON_SETTINGS = """<svg xmlns="http://www.w3.org/2000/svg" width="24" height="24" viewBox="0 0 24 24" fill="none" stroke="#f97316" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><circle cx="12" cy="12" r="3"></circle><path d="M19.4 15a1.65 1.65 0 0 0 .33 1.82l.06.06a2 2 0 0 1 0 2.83 2 2 0 0 1-2.83 0l-.06-.06a1.65 1.65 0 0 0-1.82-.33 1.65 1.65 0 0 0-1 1.51V21a2 2 0 0 1-2 2 2 2 0 0 1-2-2v-.09A1.65 1.65 0 0 0 9 19.4a1.65 1.65 0 0 0-1.82.33l-.06.06a2 2 0 0 1-2.83 0 2 2 0 0 1 0-2.83l.06-.06a1.65 1.65 0 0 0 .33-1.82 1.65 1.65 0 0 0-1.51-1H3a2 2 0 0 1-2-2 2 2 0 0 1 2-2h.09A1.65 1.65 0 0 0 4.6 9a1.65 1.65 0 0 0-.33-1.82l-.06-.06a2 2 0 0 1 0-2.83 2 2 0 0 1 2.83 0l.06.06a1.65 1.65 0 0 0 1.82.33H9a1.65 1.65 0 0 0 1-1.51V3a2 2 0 0 1 2-2 2 2 0 0 1 2 2v.09a1.65 1.65 0 0 0 1 1.51 1.65 1.65 0 0 0 1.82-.33l.06-.06a2 2 0 0 1 2.83 0 2 2 0 0 1 0 2.83l-.06.06a1.65 1.65 0 0 0-.33 1.82V9a1.65 1.65 0 0 0 1.51 1H21a2 2 0 0 1 2 2 2 2 0 0 1-2 2h-.09a1.65 1.65 0 0 0-1.51 1z"></path></svg>"""
ICON_MAIN = """<svg xmlns="http://www.w3.org/2000/svg" width="32" height="32" viewBox="0 0 24 24" fill="none" stroke="#ea580c" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M3 9l9-7 9 7v11a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2z"></path><polyline points="9 22 9 12 15 12 15 22"></polyline></svg>"""

# --- Utility Functions ---

@st.cache_data
def load_master_csv(base_path, file_pattern):
    """Load master CSV from assets directory."""
    search_path = os.path.join(base_path, f'*{file_pattern}*.csv')
    list_of_files = glob.glob(search_path)
    if not list_of_files:
        return pd.DataFrame(), None
    latest_file = max(list_of_files, key=os.path.getmtime)
    encodings = ['utf-8-sig', 'utf-8', 'cp932', 'shift_jis']
    for encoding in encodings:
        try:
            df = pd.read_csv(latest_file, encoding=encoding, dtype=str).fillna('')
            if not df.empty:
                df.columns = df.columns.str.strip()
                return df, os.path.basename(latest_file)
        except Exception:
            continue
    return pd.DataFrame(), None

def save_master_file(base_path, uploaded_file, file_pattern):
    """Save uploaded master file to assets directory, removing old ones."""
    # 1. Delete existing files matching the pattern
    search_path = os.path.join(base_path, f'*{file_pattern}*.csv')
    old_files = glob.glob(search_path)
    for f in old_files:
        try:
            os.remove(f)
        except Exception as e:
            st.error(f"旧ファイルの削除に失敗しました: {str(e)}")
            return False

    # 2. Save new file
    save_path = os.path.join(base_path, uploaded_file.name)
    try:
        with open(save_path, "wb") as f:
            f.write(uploaded_file.getbuffer())
        return True
    except Exception as e:
        st.error(f"ファイルの保存に失敗しました: {str(e)}")
        return False

def clear_sheet(ws):
    """Clear all cells in a worksheet."""
    if ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)

# --- Main App Logic ---

st.markdown(f'<div class="main-header">{ICON_MAIN} ママミール業務ツール</div>', unsafe_allow_html=True)

# Sidebar
with st.sidebar:
    st.header("設定")
    # Streamlit Cloud uses st.secrets, local uses .env
    api_key = None
    try:
        # Try Streamlit Cloud secrets first
        api_key = st.secrets.get("GOOGLE_API_KEY")
    except:
        # Fall back to environment variable
        api_key = os.environ.get("GOOGLE_API_KEY")
    
    if api_key:
        st.success("API Key: 設定済み")
        genai.configure(api_key=api_key)
    else:
        st.error("API Key: 未設定 (Streamlit Cloudの場合はSecretsを、ローカルの場合は.envを確認してください)")
    
    model_name = st.selectbox(
        "使用モデル (シール作成用)",
        ["gemini-2.5-pro", "gemini-2.5-flash", "gemini-2.0-flash"],
        index=1,
        help="gemini-2.5-pro: 高精度ですが時間がかかります\ngemini-2.5-flash: 高速ですが精度が落ちる場合があります"
    )

ASSETS_DIR = APP_DIR / 'api' / 'assets'
if not ASSETS_DIR.exists():
    ASSETS_DIR.mkdir(parents=True, exist_ok=True)
ASSETS_DIR = str(ASSETS_DIR)  # Convert to string for compatibility with glob and other functions

# Initialize Session State
if 'main_process_done' not in st.session_state:
    st.session_state.main_process_done = False
    st.session_state.template_bytes = None
    st.session_state.nouhinsyo_bytes = None
    st.session_state.original_filename = ""

if 'seal_process_done' not in st.session_state:
    st.session_state.seal_process_done = False
    st.session_state.seal_bytes = None
    st.session_state.seal_filename = ""
    st.session_state.seal_blocks = []

# Tabs
tab1, tab2, tab3 = st.tabs(["📄 数出表・納品書作成", "🏷️ シール作成", "⚙️ マスタ管理"])

# --- Tab 1: Order/Invoice Processing (Rule-based) ---
with tab1:
    st.markdown(f'<div class="sub-header">{ICON_DOC} 数出表・納品書作成</div>', unsafe_allow_html=True)
    st.markdown('<div class="card">PDFから情報を抽出し、数出表と納品書を作成します。</div>', unsafe_allow_html=True)
    
    uploaded_file_order = st.file_uploader("注文PDFをアップロード", type=['pdf'], key="order_pdf")
    
    if uploaded_file_order:
        if st.button("変換開始", key="btn_order"):
            try:
                with st.spinner('PDFを解析中...'):
                    # Load Masters
                    df_product_master, _ = load_master_csv(ASSETS_DIR, "商品マスタ")
                    df_customer_master, _ = load_master_csv(ASSETS_DIR, "得意先マスタ")
                    
                    pdf_bytes_io = io.BytesIO(uploaded_file_order.getvalue())
                    original_pdf_name = os.path.splitext(uploaded_file_order.name)[0]
                    
                    # Load Templates
                    template_path = os.path.join(ASSETS_DIR, "template.xlsm")
                    nouhinsyo_path = os.path.join(ASSETS_DIR, "nouhinsyo.xlsx")
                    
                    if not os.path.exists(template_path) or not os.path.exists(nouhinsyo_path):
                        st.error("テンプレートファイルが見つかりません。")
                        st.stop()

                    template_wb = load_workbook(template_path, keep_vba=True)
                    nouhinsyo_wb = load_workbook(nouhinsyo_path)
                    
                    # 1. Paste Masters to Template (Clear & Paste)
                    if not df_product_master.empty and "商品マスタ" in template_wb.sheetnames:
                        ws = template_wb["商品マスタ"]
                        clear_sheet(ws) # Clear existing
                        paste_dataframe_to_sheet(ws, df_product_master)
                        
                    if not df_customer_master.empty and "得意先マスタ" in template_wb.sheetnames:
                        ws = template_wb["得意先マスタ"]
                        clear_sheet(ws) # Clear existing
                        paste_dataframe_to_sheet(ws, df_customer_master)

                    # 2. Extract Data from PDF (Rule-based)
                    df_paste_sheet = pdf_to_excel_data_for_paste_sheet(io.BytesIO(pdf_bytes_io.getvalue()))
                    
                    if df_paste_sheet is None:
                        st.error("PDFデータの抽出に失敗しました。")
                        st.stop()
                        
                    # Extract Bento Data
                    df_bento_sheet = None
                    tables = extract_table_from_pdf_for_bento(io.BytesIO(pdf_bytes_io.getvalue()))
                    if tables:
                        main_table = max(tables, key=len)
                        anchor_col = find_correct_anchor_for_bento(main_table)
                        if anchor_col != -1:
                            bento_list = extract_bento_range_for_bento(main_table, anchor_col)
                            if bento_list:
                                matched_data = match_bento_data(bento_list, df_product_master)
                                df_bento_sheet = pd.DataFrame(matched_data, columns=['商品予定名', 'パン箱入数', '売価単価', '弁当区分'])
                    
                    # Extract Client Data
                    df_client_sheet = None
                    client_data = extract_detailed_client_info_from_pdf(io.BytesIO(pdf_bytes_io.getvalue()))
                    if client_data:
                        df_client_sheet = export_detailed_client_data_to_dataframe(client_data)
                        
                    # 3. Write to Template
                    ws_paste = template_wb["貼り付け用"]
                    for r_idx, row in df_paste_sheet.iterrows():
                        for c_idx, value in enumerate(row):
                            ws_paste.cell(row=r_idx + 1, column=c_idx + 1, value=value)
                            
                    if df_bento_sheet is not None and "注文弁当の抽出" in template_wb.sheetnames:
                        safe_write_df(template_wb["注文弁当の抽出"], df_bento_sheet)
                        
                    if df_client_sheet is not None and "クライアント抽出" in template_wb.sheetnames:
                        safe_write_df(template_wb["クライアント抽出"], df_client_sheet)
                        
                    # 4. Write to Nouhinsyo
                    # Paste Masters to Nouhinsyo (Clear & Paste)
                    if not df_customer_master.empty and "得意先マスタ" in nouhinsyo_wb.sheetnames:
                        ws = nouhinsyo_wb["得意先マスタ"]
                        clear_sheet(ws) # Clear existing
                        paste_dataframe_to_sheet(ws, df_customer_master)
                    
                    # Note: Nouhinsyo might not have "商品マスタ" sheet, but if it does, we should update it too.
                    # Based on previous code, it only updated "得意先マスタ".
                    
                    ws_paste_n = nouhinsyo_wb["貼り付け用"]
                    for r_idx, row in df_paste_sheet.iterrows():
                        for c_idx, value in enumerate(row):
                            ws_paste_n.cell(row=r_idx + 1, column=c_idx + 1, value=value)
                            
                    # Map Bento Names for Nouhinsyo
                    df_bento_for_nouhin = None
                    if df_bento_sheet is not None:
                        master_df = df_product_master.copy()
                        if not master_df.empty and '商品名' in master_df.columns:
                            master_map = master_df.drop_duplicates(subset=['商品予定名']).set_index('商品予定名')['商品名'].to_dict()
                            df_bento_for_nouhin = df_bento_sheet.copy()
                            df_bento_for_nouhin['商品名'] = df_bento_for_nouhin['商品予定名'].map(master_map)
                            df_bento_for_nouhin = df_bento_for_nouhin[['商品予定名', 'パン箱入数', '商品名']]
                            
                    if df_bento_for_nouhin is not None and "注文弁当の抽出" in nouhinsyo_wb.sheetnames:
                        safe_write_df(nouhinsyo_wb["注文弁当の抽出"], df_bento_for_nouhin)
                        
                    if df_client_sheet is not None and "クライアント抽出" in nouhinsyo_wb.sheetnames:
                        safe_write_df(nouhinsyo_wb["クライアント抽出"], df_client_sheet)

                    # 5. Save to Session State
                    out_template = io.BytesIO()
                    template_wb.save(out_template)
                    st.session_state.template_bytes = out_template.getvalue()
                    
                    out_nouhin = io.BytesIO()
                    nouhinsyo_wb.save(out_nouhin)
                    st.session_state.nouhinsyo_bytes = out_nouhin.getvalue()
                    
                    st.session_state.original_filename = original_pdf_name
                    st.session_state.main_process_done = True
                    
                    st.success("処理が完了しました！")
                    
            except Exception as e:
                st.error(f"エラーが発生しました: {str(e)}")
                import traceback
                st.code(traceback.format_exc())

    # Display Download Buttons (Persistent)
    if st.session_state.main_process_done:
        col1, col2 = st.columns(2)
        with col1:
            st.download_button(
                label="数出表をダウンロード",
                data=st.session_state.template_bytes,
                file_name=f"{st.session_state.original_filename}_数出表.xlsm",
                mime="application/vnd.ms-excel.sheet.macroEnabled.12",
                key="dl_template"
            )
        with col2:
            st.download_button(
                label="納品書ダウンロード",
                data=st.session_state.nouhinsyo_bytes,
                file_name=f"{st.session_state.original_filename}_納品書.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_nouhin"
            )

# --- Tab 2: Seal Processing (AI) ---
with tab2:
    st.markdown(f'<div class="sub-header">{ICON_TAG} シール作成</div>', unsafe_allow_html=True)
    st.markdown('<div class="card">PDFからシール用データを抽出し、Excelを作成します。</div>', unsafe_allow_html=True)
    
    uploaded_file_seal = st.file_uploader("シールPDFをアップロード", type=['pdf'], key="seal_pdf")
    
    if uploaded_file_seal:
        if st.button("変換開始", key="btn_seal"):
            try:
                with st.spinner('AIが解析中... これには数分かかる場合があります。'):
                    model = genai.GenerativeModel(model_name)
                    pdf_bytes = uploaded_file_seal.getvalue()
                    
                    seal_prompt = """
このPDFはシール表です。横4つ × 縦5つ(合計約20個)のブロックで構成されています。
各ブロックには以下の情報が含まれています:
1. クライアント名 (最上部): 小学校名または幼稚園名 + 「様」
2. 準備物 (クライアント名のすぐ下): パン箱入数、ご飯150gなど
3. クラス名 (中央、大きめの文字): チューリップ、さくらなど
4. 弁当数 (クラス名の下): 数値(例: 35、35+1)
5. 日付 (ブロック左下): MM/DD形式
6. 学年 (ブロック右下): 年長、年中など

以下のJSON形式で、全てのブロック情報を抽出してください:
{
  "blocks": [
    {
      "client_name": "博多南衆参コース様",
      "preparations": ["パン箱入数", "ご飯150g"],
      "class_name": "チューリップ",
      "meal_count": "35",
      "date": "12/10",
      "grade": "年長"
    }
  ]
}
重要: 全てのブロックを抽出してください。完全で有効なJSONのみを返してください。
"""
                    response = model.generate_content([
                        {"mime_type": "application/pdf", "data": pdf_bytes},
                        seal_prompt
                    ], generation_config={"response_mime_type": "application/json"})
                    
                    text = response.text.strip()
                    if text.startswith("```json"): text = text[7:]
                    if text.endswith("```"): text = text[:-3]
                    data = json.loads(text)
                    
                    blocks = data if isinstance(data, list) else data.get('blocks', [])
                    
                    # Create Excel
                    seal_path = os.path.join(ASSETS_DIR, "seal.xlsx")
                    if os.path.exists(seal_path):
                        wb = load_workbook(seal_path)
                    else:
                        wb = Workbook()
                    
                    target_sheet_name = "Gemini抽出データ"
                    if target_sheet_name in wb.sheetnames:
                        ws = wb[target_sheet_name]
                        ws.delete_rows(1, ws.max_row + 1)
                    else:
                        ws = wb.create_sheet(title=target_sheet_name)
                    
                    wb.active = ws
                    headers = ['クライアント名', 'クラス名', '準備物', '弁当数', '日付', '学年']
                    ws.append(headers)
                    
                    for block in blocks:
                        prep = block.get('preparations', [])
                        prep_text = ', '.join(prep) if isinstance(prep, list) else str(prep)
                        ws.append([
                            block.get('client_name', ''),
                            block.get('class_name', ''),
                            prep_text,
                            block.get('meal_count', ''),
                            block.get('date', ''),
                            block.get('grade', '')
                        ])
                    
                    out_seal = io.BytesIO()
                    wb.save(out_seal)
                    
                    st.session_state.seal_bytes = out_seal.getvalue()
                    st.session_state.seal_filename = uploaded_file_seal.name.replace('.pdf', '') + '_seal.xlsx'
                    st.session_state.seal_blocks = blocks
                    st.session_state.seal_process_done = True
                    
                    st.success("処理が完了しました！")
                        
            except Exception as e:
                st.error(f"エラーが発生しました: {str(e)}")
                import traceback
                st.code(traceback.format_exc())

    if st.session_state.seal_process_done:
        st.write(f"抽出されたデータ数: {len(st.session_state.seal_blocks)}件")
        st.download_button(
            label="シールデータ (Excel) をダウンロード",
            data=st.session_state.seal_bytes,
            file_name=st.session_state.seal_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_seal"
        )
        with st.expander("抽出データプレビュー"):
            st.json(st.session_state.seal_blocks)

# --- Tab 3: Master Management ---
with tab3:
    st.markdown(f'<div class="sub-header">{ICON_SETTINGS} マスタ管理</div>', unsafe_allow_html=True)
    st.markdown('<div class="card">商品マスタと得意先マスタのCSVをアップロードします。</div>', unsafe_allow_html=True)
    
    # Show current files
    st.subheader("現在のマスタファイル")
    _, prod_file = load_master_csv(ASSETS_DIR, "商品マスタ")
    _, cust_file = load_master_csv(ASSETS_DIR, "得意先マスタ")
    
    col1, col2 = st.columns(2)
    with col1:
        st.info(f"商品マスタ: {prod_file if prod_file else 'なし'}")
    with col2:
        st.info(f"得意先マスタ: {cust_file if cust_file else 'なし'}")

    # Uploaders
    st.subheader("マスタ更新")
    st.markdown("※アップロードすると、古いマスタファイルは削除され、新しいファイルが保存されます。")
    
    up_prod = st.file_uploader("商品マスタをアップロード (CSV)", type=['csv'], key="up_prod")
    if up_prod:
        if "商品マスタ一覧" not in up_prod.name:
            st.warning("ファイル名に「商品マスタ一覧」を含めてください。")
        else:
            if st.button("商品マスタを保存"):
                if save_master_file(ASSETS_DIR, up_prod, "商品マスタ"):
                    st.success(f"保存しました: {up_prod.name}")
                    st.rerun()

    up_cust = st.file_uploader("得意先マスタをアップロード (CSV)", type=['csv'], key="up_cust")
    if up_cust:
        if "得意先マスタ一覧" not in up_cust.name:
            st.warning("ファイル名に「得意先マスタ一覧」を含めてください。")
        else:
            if st.button("得意先マスタを保存"):
                if save_master_file(ASSETS_DIR, up_cust, "得意先マスタ"):
                    st.success(f"保存しました: {up_cust.name}")
                    st.rerun()
